from tkinter import *
from openpyxl import load_workbook
import tkinter.ttk as ttk

class MyWindow:
        
    def __init__(self, win):

        win.title("Carbon Footprint Calculator")

        lb2= Label(win, text="Carbon Footprint Calculator", font=("arial",14))  
        lb2.place(x=160,y=40) 


        wb = load_workbook(filename=r"C:\Users\hardik\Downloads\testpy.xlsx")
        ws = wb['Sheet1']
        xlsx_range = ws['B2':'B7']
        elements = []

        for cell in xlsx_range:
            for x in cell:
                y = x.value
                elements.append(y)
                print(y)      
        lb2= Label(win, text="States", width=13,font=("arial",12))  
        lb2.place(x=14,y=300)  
        combodata = ttk.Combobox(win, state='readonly', values=elements)
        combodata.place(x=200, y=80)

        '''list_of_cntry1 = ("Smart Mobiles", "smartwatches", "smart fire alarms", "smart door locks","smart bicycles","medical sensors")  
        cv = StringVar()  
        drplist= OptionMenu(win, cv, *list_of_cntry1)  
        drplist.config(width=16)  
        cv.set("Select Smart Product")'''
        lb2= Label(win, text="Smart Product", font=("arial",12))  
        lb2.place(x=14,y=80)  
        #drplist.place(x=200, y=75)

        list_of_cntry2 = (1,2,3,4,5,6)  
        cv = StringVar()  
        drplist= OptionMenu(win, cv, *list_of_cntry2)  
        drplist.config(width=15)  
        cv.set("Select No.Rack")  
        lb1= Label(win, text="Number of Racks", width=13,font=("arial",12))  
        lb1.place(x=14,y=120)  
        drplist.place(x=200, y=120)

        self.lbl1=Label(win, text='Total heat load(TR)',  font=("arial",12))
        self.lbl3=Label(win, text='Total Electronic Load(kW)',  font=("arial",12))
        self.t1=Entry(bd=3)       
        self.t3=Entry()       
        
        self.btn3=Button(win,text="Calculate Bill")
        
        self.lbl1.place(x=14, y=150)
        self.t1.place(x=200, y=150)    
               
        self.b3=Button(win,text="Calculate Bill")
        self.b3.bind('<Button-1>',self.mul)        
        
        self.b3.place(x=300,y=150)       
        self.lbl3.place(x=14, y=200)
        self.t3.place(x=200, y=200)
     

        '''list_of_cntry = ("Maharashtra", "Delhi", "Uttar Pradesh", "Madhya Pradesh","Bihar","Goa")  
        cv = StringVar()  
        drplist= OptionMenu(win, cv, *list_of_cntry)  
        drplist.config(width=16)  
        cv.set("State List")  
        lb2= Label(win, text="States", width=13,font=("arial",12))  
        lb2.place(x=14,y=300)  
        drplist.place(x=200, y=300)'''

        self.b4=Button(win,text="State Bill Calculate")
        self.b4.bind('<Button-1>',self.mul2)   
        self.b4.place(x=380,y=350)

        self.lbl6=Label(win, text='Tariff Rate',  font=("arial",12))
        self.lbl6.place(x=14, y=325)
        


        self.lbl4=Label(win, text='Carbon footprint Calculator',  font=("arial",12))
        self.lbl4.place(x=14, y=350)

        self.t5=Entry()      
        self.t5.place(x=220, y=350)

        wb = load_workbook(filename=r"C:\Users\hardik\Downloads\testpy.xlsx")
        ws = wb['Sheet1']
        xlsx_range = ws['A2':'A29']
        elements = []

        for cell in xlsx_range:
            for x in cell:
                y = x.value
                elements.append(y)
                print(y)      
        lb2= Label(win, text="States", width=13,font=("arial",12))  
        lb2.place(x=14,y=300)  
        combodata = ttk.Combobox(win, state='readonly', values=elements)
        combodata.place(x=200, y=300)


        
    def mul(self, event):
        self.t3.delete(0, 'end')
        num1=int(self.t1.get())
        num2=float(3.54)
        result=num1*num2
        self.t3.insert(END, str(result))
    def mul2(self, event):
        self.t5.delete(0, 'end')
        num1=float(self.t3.get())
        num2=float(0.86)
        result=num1*num2
        self.t5.insert(END, str(result))

window=Tk()
mywin=MyWindow(window)
window.title('Carbon Footprint Calculator')
window.geometry("700x500+10+10")
window.mainloop()